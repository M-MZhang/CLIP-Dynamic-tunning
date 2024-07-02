"""
Goal
---
1. Read test results from log.txt files
2. Compute mean and std across different folders (seeds)

Usage
---
Assume the output files are saved under output/my_experiment,
which contains results of different seeds, e.g.,

my_experiment/
    seed1/
        log.txt
    seed2/
        log.txt
    seed3/
        log.txt

Run the following command from the root directory:

$ python tools/parse_test_res.py output/my_experiment

Add --ci95 to the argument if you wanna get 95% confidence
interval instead of standard deviation:

$ python tools/parse_test_res.py output/my_experiment --ci95

If my_experiment/ has the following structure,

my_experiment/
    exp-1/
        seed1/
            log.txt
            ...
        seed2/
            log.txt
            ...
        seed3/
            log.txt
            ...
    exp-2/
        ...
    exp-3/
        ...

Run

$ python tools/parse_test_res.py output/my_experiment --multi-exp
"""
import re
import numpy as np
import os.path as osp
import argparse
import openpyxl
import os
from collections import OrderedDict, defaultdict

from dassl.utils import check_isfile, listdir_nohidden


def compute_ci95(res):
    return 1.96 * np.std(res) / np.sqrt(len(res))


def parse_function(*metrics, directory="", args=None, end_signal=None):
    print(f"Parsing files in {directory}")
    subdirs = listdir_nohidden(directory, sort=True)

    outputs = []

    for subdir in subdirs:
        fpath = osp.join(directory, subdir, "log.txt")
        assert check_isfile(fpath)
        good_to_go = False
        output = OrderedDict()

        with open(fpath, "r") as f:
            lines = f.readlines()

            for line in lines:
                line = line.strip()

                if line == end_signal:
                    good_to_go = True

                for metric in metrics:
                    match = metric["regex"].search(line)
                    if match and good_to_go:
                        if match:
                            if "file" not in output:
                                output["file"] = fpath
                            num = float(match.group(1))
                            name = metric["name"]
                            output[name] = num

        if output:
            outputs.append(output)

    assert len(outputs) > 0, f"Nothing found in {directory}"

    metrics_results = defaultdict(list)

    for output in outputs:
        msg = ""
        for key, value in output.items():
            if isinstance(value, float):
                msg += f"{key}: {value:.2f}%. "
            else:
                msg += f"{key}: {value}. "
            if key != "file":
                metrics_results[key].append(value)
        print(msg)

    output_results = OrderedDict()

    print("===")
    print(f"Summary of directory: {directory}")
    for key, values in metrics_results.items():
        avg = np.mean(values)
        std = compute_ci95(values) if args.ci95 else np.std(values)
        print(f"* {key}: {avg:.2f}% +- {std:.2f}%")
        output_results[key] = avg
    print("===")

    return output_results, std


def main(args, end_signal):
    # for train-test results
    metric = {
        "name": args.keyword,
        "regex": re.compile(fr"\* {args.keyword}: ([\.\deE+-]+)%"),
    }

    # for time-test results
    # metric = {
    #     "name": args.keyword,
    #     "regex": re.compile(fr"\* {args.keyword}: ([\.\deE+-]+) im/s"),
    # }

    if args.multi_exp:
        final_results = defaultdict(list)

        for directory in listdir_nohidden(args.directory, sort=True):
            directory = osp.join(args.directory, directory)
            results, stds = parse_function(
                metric, directory=directory, args=args, end_signal=end_signal
            )

            for key, value in results.items():
                final_results[key].append(value)

        print("Average performance")
        for key, values in final_results.items():
            avg = np.mean(values)
            print(f"* {key}: {avg:.2f}%")

    else:
        results, stds = parse_function(
            metric, directory=args.directory, args=args, end_signal=end_signal
        )
        return results, stds


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--directory", type=str, default=None, help="path to directory")
    parser.add_argument(
        "--ci95", action="store_true", help=r"compute 95\% confidence interval"
    )
    parser.add_argument("--test-log", action="store_true", help="parse test-only logs")
    parser.add_argument(
        "--multi-exp", action="store_true", help="parse multiple experiments"
    )
    parser.add_argument(
        "--keyword", default="accuracy", type=str, help="which keyword to extract"
    )
    args = parser.parse_args()


    # 需要写入excel表中
    save_dir = "/root/data1/zmm/output"
    save_file_name = save_dir + "/results.xlsx"
    colums_names = ["Trainer", "caltech101", "oxford_pets", "stanford_cars", "oxford_flowers", "food101", "fgvc_aircraft", "sun397", "dtd", "eurosat", "ucf101"]

    if not os.path.exists(save_file_name):
        wb = openpyxl.Workbook(save_file_name)
        sheet = wb.create_sheet('Comparision-experiments-new') # 根据trainer不同需要替换
        sheet.append(colums_names)
        wb.save(save_file_name)
    
    train_average_list = []
    test_average_list = []
    tome_test_average_list = []
    for dataset in colums_names[1:]:
        
        # end_signal = "Finish training"
        # train_dictionary = "/root/data1/zmm/output/base2new/train_base/" + dataset + "/shots_16/ReMaPLe_4/vit_b16_c2_ep5_batch4_2ctx"
        # args.directory = train_dictionary
        # results, stds = main(args, end_signal)
        # train_average_list.append(str(round(results['accuracy'],2)) + " % +- " + str(round(stds,2)) + " %")

        
        test_dictionary = "/root/data1/zmm/output/base2new/test_new/" + dataset + "/shots_16/ZeroshotCLIP_ToMe/vit_b16"
        args.directory = test_dictionary
        # args.test_log:
        end_signal = "=> result"
        # end_signal = "Throughput"
        results, stds = main(args, end_signal=end_signal)
        # test_average_list.append(str(round(results['Throughput'],2)) + " % +-" + str(round(stds,2)) + " %")
        test_average_list.append(str(round(results['accuracy'],2)) + " % +- " + str(round(stds,2)) + " %")

        # tome_test_dictionary = "/root/data1/zmm/output/base2new/test_new/" + dataset + "/shots_16/ReMaPLe_ToMe/vit_b16_c2_ep5_batch4_2ctx"
        # args.directory = tome_test_dictionary
        # # args.test_log:
        # end_signal = "=> result"
        # results, stds = main(args, end_signal=end_signal)
        # tome_test_average_list.append(str(round(results['accuracy'],2)) + " % +- " + str(round(stds,2)) + " %")
    

    wb = openpyxl.load_workbook(save_file_name)
    if "Comparision-experiments-new" in wb.sheetnames:
        sheet = wb['Comparision-experiments-new']
    else:
        sheet = wb.create_sheet('Comparision-experiments-new')
        sheet = wb['Comparision-experiments-new']
        sheet.append(colums_names)
    
    # change name for different trainers
    # train_average_list = ["ReMaPLe_4_base"] + train_average_list
    test_average_list = ["ZeroshotCLIP_ToMe"] + test_average_list
    # tome_test_average_list = ["MaPLe_ToMe_novel"] + tome_test_average_list
    # sheet.append(train_average_list)
    sheet.append(test_average_list)
    # sheet.append(tome_test_average_list)
    wb.save(filename=save_file_name)